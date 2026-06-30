"""
excel_parser.py
----------------
Parses Screener.in's "Export to Excel" file into the same flat dict shape
that data_fetcher.load_sample() / fetch_live() already return -- so it
plugs straight into the existing app.py -> quant_engine.py pipeline with
zero changes to the calculation logic.

Why this approach: Screener's export always contains a "Data Sheet" tab
in a fixed template (the file itself declares 'LATEST VERSION 2.1').
The Profit & Loss / Balance Sheet / Cash Flow / Price blocks always use
the same row labels in the same order, regardless of company -- only the
number of optional expense sub-lines (Raw Material Cost, Power and Fuel,
etc.) varies. That's what makes this generic instead of TCS-specific.

This deliberately reads from official annual report figures (via
Screener, which sources from BSE/NSE filings) rather than yfinance --
this is the accurate, audited data source the user asked for, replacing
the live-scrape pipeline that kept producing wrong/zero figures.
"""

from __future__ import annotations
import openpyxl


SECTION_HEADERS = {"META", "PROFIT & LOSS", "Quarters", "BALANCE SHEET", "CASH FLOW:", "PRICE:", "DERIVED:"}


def _find_section_rows(ws, start_label: str, end_label: str | None):
    """Returns {row_label: [values...]} for every row strictly between the
    row whose col-A text == start_label and the row whose col-A text ==
    end_label (or end of sheet if end_label is None)."""
    rows = list(ws.iter_rows(values_only=True))
    start_idx = end_idx = None
    for i, row in enumerate(rows):
        label = row[0]
        if isinstance(label, str) and label.strip() == start_label:
            start_idx = i
        elif start_idx is not None and end_label and isinstance(label, str) and label.strip() == end_label:
            end_idx = i
            break
    if start_idx is None:
        return {}
    if end_idx is None:
        end_idx = len(rows)

    section = {}
    for row in rows[start_idx + 1:end_idx]:
        label = row[0]
        if isinstance(label, str) and label.strip():
            section[label.strip()] = list(row[1:])
    return section


def _row(section: dict, *label_options, n: int) -> list[float]:
    """Looks up a row by trying several possible label spellings (Screener
    is mostly consistent but a few labels differ slightly, e.g. 'Cash &
    Bank' vs 'Cash and Bank Balance'). Missing values default to 0 rather
    than crashing -- matches the same safe-default philosophy already used
    in quant_engine.py."""
    for label in label_options:
        if label in section:
            vals = section[label][:n]
            vals = vals + [0] * (n - len(vals))
            return [float(v) if isinstance(v, (int, float)) else 0.0 for v in vals]
    return [0.0] * n


def parse_screener_excel(file) -> dict:
    """file: a path string, or a file-like object (e.g. Streamlit's
    UploadedFile from st.file_uploader -- openpyxl accepts both)."""
    wb = openpyxl.load_workbook(file, data_only=True)
    if "Data Sheet" not in wb.sheetnames:
        raise ValueError(
            "This doesn't look like a Screener.in export -- expected a 'Data Sheet' "
            "tab. Use the 'Export to Excel' button on the company's Screener page."
        )
    ws = wb["Data Sheet"]

    # ---- META ----
    meta = _find_section_rows(ws, "META", "PROFIT & LOSS")
    company_name = None
    for row in ws.iter_rows(values_only=True):
        if row[0] == "COMPANY NAME":
            company_name = row[1]
            break
    current_price = (meta.get("Current Price") or [0])[0] or 0
    face_value = (meta.get("Face Value") or [10])[0] or 10

    # ---- ANNUAL PROFIT & LOSS ----
    pl = _find_section_rows(ws, "PROFIT & LOSS", "Quarters")
    dates = pl.get("Report Date", [])
    years = [d.year for d in dates if d is not None]
    n = len(years)
    if n == 0:
        raise ValueError("Could not find annual report dates in the 'PROFIT & LOSS' section of this file.")

    sales = _row(pl, "Sales", n=n)
    raw_material = _row(pl, "Raw Material Cost", n=n)
    change_inv = _row(pl, "Change in Inventory", n=n)
    power_fuel = _row(pl, "Power and Fuel", n=n)
    other_mfr = _row(pl, "Other Mfr. Exp", n=n)
    employee_cost = _row(pl, "Employee Cost", n=n)
    selling_admin = _row(pl, "Selling and admin", n=n)
    other_expenses = _row(pl, "Other Expenses", n=n)
    other_income = _row(pl, "Other Income", n=n)
    depreciation = _row(pl, "Depreciation", n=n)
    interest = _row(pl, "Interest", n=n)
    pbt = _row(pl, "Profit before tax", n=n)
    tax = _row(pl, "Tax", n=n)
    net_profit = _row(pl, "Net profit", n=n)
    dividend_amount = _row(pl, "Dividend Amount", n=n)

    # Screener's annual P&L doesn't give a direct "Operating Profit" line
    # (unlike its Quarters section) -- it's Sales minus the cost lines.
    operating_profit = [
        s - (rm + ci + pf + om + ec + sa + oe)
        for s, rm, ci, pf, om, ec, sa, oe in zip(
            sales, raw_material, change_inv, power_fuel, other_mfr,
            employee_cost, selling_admin, other_expenses
        )
    ]

    # ---- BALANCE SHEET ----
    bs = _find_section_rows(ws, "BALANCE SHEET", "CASH FLOW:")
    equity_capital = _row(bs, "Equity Share Capital", n=n)
    reserves = _row(bs, "Reserves", n=n)
    borrowings = _row(bs, "Borrowings", n=n)
    other_liabilities = _row(bs, "Other Liabilities", n=n)
    net_block = _row(bs, "Net Block", n=n)
    investments = _row(bs, "Investments", n=n)
    debtors = _row(bs, "Receivables", "Debtors", n=n)
    inventory = _row(bs, "Inventory", n=n)
    cash = _row(bs, "Cash & Bank", "Cash and Bank Balance", n=n)

    # ---- CASH FLOW ----
    cf = _find_section_rows(ws, "CASH FLOW:", "PRICE:")
    cfo = _row(cf, "Cash from Operating Activity", n=n)

    # Screener doesn't export a direct Capex line. Approximate it from the
    # balance sheet (change in Net Block + that year's Depreciation), which
    # is the standard analyst proxy when capex isn't separately reported.
    # Flagged here explicitly -- it's an approximation, not a reported figure.
    capex = [0.0] * n
    for i in range(n):
        prev_block = net_block[i - 1] if i > 0 else net_block[i]
        capex[i] = max((net_block[i] - prev_block) + depreciation[i], 0.0)

    # ---- PRICE row (label + values on the same row) ----
    price = [current_price] * n
    for row in ws.iter_rows(values_only=True):
        if row[0] == "PRICE:":
            vals = list(row[1:1 + n])
            price = [float(v) if isinstance(v, (int, float)) else current_price for v in vals]
            break

    # ---- DERIVED: Adjusted Equity Shares in Cr (needed for accurate EPS) ----
    shares_cr = [None] * n
    derived = _find_section_rows(ws, "DERIVED:", None)
    if "Adjusted Equity Shares in Cr" in derived:
        vals = derived["Adjusted Equity Shares in Cr"][:n]
        shares_cr = [float(v) if isinstance(v, (int, float)) else None for v in vals]

    eps = []
    for i in range(n):
        if shares_cr[i]:
            eps.append(round(net_profit[i] / shares_cr[i], 2))
        else:
            eps.append(0.0)

    dividend_payout = [
        round(da / np_, 4) if np_ else 0.0 for da, np_ in zip(dividend_amount, net_profit)
    ]

    return {
        "company": company_name or "Unknown Company",
        "sector": "Unknown Sector",  # Screener's export doesn't include sector/industry
        "face_value": face_value,
        "current_price": current_price,
        "govt_bond_yield": 0.071,
        "years": years,
        "sales": sales,
        "operating_profit": operating_profit,
        "other_income": other_income,
        "interest": interest,
        "pbt": pbt,
        "tax": tax,
        "net_profit": net_profit,
        "eps": eps,
        "price": price,
        "dividend_payout": dividend_payout,
        "equity_capital": equity_capital,
        "reserves": reserves,
        "borrowings": borrowings,
        "other_liabilities": other_liabilities,
        "net_block": net_block,
        "investments": investments,
        "debtors": debtors,
        "inventory": inventory,
        "cash": cash,
        "cfo": cfo,
        "capex": capex,  # NOTE: approximated -- see comment above
        "governance_flags": {
            "promoter_pledge_pct": 0.0,
            "promoter_holding_change_yoy": 0.0,
            "share_dilution_pct_yoy": 0.0,
            "auditor_changed_last_5y": False,
            "related_party_transactions_flag": False,
            "results_delayed_flag": False
        }
    }
